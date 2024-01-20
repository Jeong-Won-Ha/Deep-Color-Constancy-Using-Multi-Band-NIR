import os
import torch.nn as nn
import torch.optim as optim
import numpy as np

import easydict
import socket
import time
from torch.autograd import Variable
import torchvision
from torch.utils.data import DataLoader
import torch.backends.cudnn as cudnn
from dataloader import DatasetFromFolder
# from model_ReWU import hierarchy3_confidence as Full_model
# from model import FC4 as Full_model
from model_2branch import Full_model as Full_model
from loss import *
from torchvision.transforms import *
import openpyxl

opt = easydict.EasyDict({
    "batchSize": 16,  # batch size
    "lr": 1e-2,  # learning rate
    "patch_size": 256,

    "start_iter": 1,  # 다시 돌릴때는 이거도 바꿔주기
    "nEpochs": 1000,  # training 횟수
    "snapshots": 20,  # weight 저장 주기

    "data_dir": "C:/NIR_dataset_0205/",  # dataset 저장 위치

    "model_type": "result",  # 모델이름

    "save_folder": "./weights/bright_rgb_proposed_3ch/",  # weight 저장 위치
    "resume": False,
    "pretrained": False,
    "gpu_mode": True,
    "threads": 1,
    "seed": 123,
    "gpus": 1,
    "prefix": "tpami_residual_filter8",
    # "input_dir": "H:\dataset_fabric",  # test dataset 불러올위치
    "test_dataset": "Test.pt",
    "testBatchSize": 1,

    "output": "./results/bright_rgb_proposed_3ch/",  # 결과영상 저장위치
    "testresult": "same_frame",
    "excel_name": './results/bright_rgb_proposed_3ch/est_ilu.xlsx',
    "patch_size": 256,
})

checkpoint_name2 = os.path.join(opt.save_folder + 'min_AE.pth')
excel_name = opt.excel_name

if not os.path.exists(opt.output):
    os.makedirs(opt.output)
if os.path.exists(excel_name):
    write_wb = openpyxl.load_workbook(excel_name, data_only=True)
    write_ws = write_wb.active
else:
    write_wb = openpyxl.Workbook()
    write_ws = write_wb.active


if not os.path.exists(opt.save_folder):
    os.makedirs(opt.save_folder)

def log(text, LOGGER_FILE):
    with open(LOGGER_FILE, 'a') as f:
        f.write(text)
        f.close()

gpus_list = range(opt.gpus)
hostname = str(socket.gethostname())
cudnn.benchmark = True
print(opt)

label = np.zeros(shape=(257, 3), dtype=np.float)
label2 = np.zeros(shape=(78, 3), dtype=np.float)
label[0:77,:] = [0.748821368466364, 0.567288158220369, 0.342710816390589]
label[77:156,:] = [0.794080769118038, 0.538175815166851, 0.282493405385683]
label[156:233,:] = [0.834806408654636, 0.505058216168657, 0.219122017035446]
label[233:241,:] = [0.833468706405267, 0.506333341309169, 0.221261074122458]
label[241:249,:] = [0.79648869429934, 0.536228824980606, 0.279400084311461]
label[249:257,:] = [0.755398531287944, 0.563615487844529, 0.334231418017203]

label2[0:23,:] = [0.748821368466364, 0.567288158220369, 0.342710816390589]
label2[23:46,:] = [0.794080769118038, 0.538175815166851, 0.282493405385683]
label2[46:69,:] = [0.834806408654636, 0.505058216168657, 0.219122017035446]
label2[69:72,:] = [0.833468706405267, 0.506333341309169, 0.221261074122458]
label2[72:75,:] = [0.79648869429934, 0.536228824980606, 0.279400084311461]
label2[75:78,:] = [0.755398531287944, 0.563615487844529, 0.334231418017203]

def eval(epoch):
    avg_angle = 0
    count = 1
    model.eval()
    for batch in testing_data_loader:
        with torch.no_grad():
            input,nir, gt, name = Variable(batch[0]), Variable(batch[1]), Variable(batch[2]), batch[3]
            # nir = nir[:,0,:,:].unsqueeze(dim=1)
        if cuda:
            input = input.cuda(gpus_list[0])
            gt = gt.cuda(gpus_list[0])
            nir = nir.cuda()
        t0 = time.time()

        with torch.no_grad():
            input = torch.cat((input,nir),dim=1)
            prediction,local,confidence = model(input)

        angle_error = criterion(prediction, gt)
        avg_angle += angle_error

        write_ws.cell(row=count, column=1).value = prediction[:, 0].item()
        write_ws.cell(row=count, column=2).value = prediction[:, 1].item()
        write_ws.cell(row=count, column=3).value = prediction[:, 2].item()
        write_ws.cell(row=count, column=4).value = angle_error.item()

        write_wb.save(excel_name)
        count = count + 1

        t1 = time.time()
        prediction = prediction / torch.norm(prediction)

        save_img((confidence).cpu().data, 'confidence_' + name[0])
        save_img((local).cpu().data, 'local_' + name[0])

        index, ext = os.path.splitext(name[0])

        print("===> Processing: %s || Timer: %.4f sec. angular error : %.4f" % ( name[0], (t1 - t0), angle_error))
        log('===> Processing: %s || Timer: %.4f sec. angular error : %.4f' % ( name, (t1 - t0), angle_error), logfile)

        average = avg_angle / len(testing_data_loader)

    print("===> Processing Done, Average Angular error : %.4f" % (average))
    log('Epoch[%d] : Test Avg loss = %.4f \n' % (epoch, average), logfile)


def checkpoint(epoch):
    model_out_path = opt.save_folder + hostname + opt.model_type + opt.prefix + "_epoch_{}.pth".format(epoch)
    torch.save(model.state_dict(), model_out_path)
    print("Checkpoint saved to {}".format(model_out_path))


def transform():
    return Compose([
        ToTensor(),
    ])

def save_img(img, img_name):
    save_dir = opt.output
    # save_dir = opt.output
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    save_fn = save_dir + '/' + img_name
    torchvision.utils.save_image(img, save_fn, normalize=True, format='png')
def save_img2(img, img_name):
    save_dir = opt.output
    # save_dir = opt.output
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    save_fn = save_dir + '/' + img_name
    torchvision.utils.save_image(img, save_fn, normalize=False, format='png')

if __name__ == '__main__':

    cuda = opt.gpu_mode
    if cuda and not torch.cuda.is_available():
        raise Exception("No GPU found, please run without --cuda")

    torch.manual_seed(opt.seed)
    if cuda:
        torch.cuda.manual_seed(opt.seed)

    print('===> Loading datasets')
    test_set = DatasetFromFolder(opt.data_dir, gt_ilu=label2, patch_size=opt.patch_size, transform=transform(),
                                 folder='test/', Num_ch='nir_3ch/', isTrain=False)
    testing_data_loader = DataLoader(dataset=test_set, num_workers=opt.threads, batch_size=opt.testBatchSize,
                                     shuffle=False)

    print('===> Building model ', opt.model_type)

    model = Full_model(in_ch=3+3)

    model = torch.nn.DataParallel(model, device_ids=gpus_list)
    criterion = angular_loss

    if cuda:
        model = model.cuda()
        # criterion = criterion.cuda(gpus_list[0])
    logfile = opt.output + 'result.txt'


    print('==> Resuming from checkpoint..')
    checkpoint_load = torch.load(checkpoint_name2)
    model.load_state_dict(checkpoint_load)
    start_epoch = 100

    eval(start_epoch)

